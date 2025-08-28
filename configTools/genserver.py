from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import exceltool
import template
import genbase
import os
import shutil
import re
import sys
import time




def get_excel_files(folder_path):
    excel_files = []
    for root, dirs, files in os.walk(folder_path):
        for filename in files:
            file_path = os.path.join(root, filename)
            if filename.lower().endswith(".xlsx") or filename.lower().endswith(".xls"):
                excel_files.append(file_path)
    return excel_files


    
script_path = os.path.abspath(__file__)
dir_path = os.path.dirname(script_path)
parent_dir = os.path.dirname(dir_path)
#打包工具需要
#parent_dir = os.path.dirname(parent_dir)
#dir_path = os.path.dirname(parent_dir)
#parent_dir = os.path.dirname(parent_dir)
#parent_dir = os.path.dirname(parent_dir)
       

lubanOutputPath = dir_path + r"\Luban\ConfigRoot\Datas"

my_excel = []
folder_paths = []

configDir = parent_dir + r"\config\JuiceZombies"


folder_paths.append(configDir)




def gen_base(isClient):


    # 获取所有文件夹下的 XLSX 和 XLS 文件路径
    excel_files = []
    for folder_path in folder_paths:
        excel_files.extend(get_excel_files(folder_path))

    my_excel.extend(excel_files)

    tablespath = lubanOutputPath + r"\__tables__.xlsx"
    workbook = load_workbook(tablespath)  # 替换为实际的文件名
    worksheet = workbook['Sheet1']  # 替换为实际的工作表名
    worksheet.delete_rows(2, worksheet.max_row)
    workbook.save(tablespath)
    workbook.close()
    #print('begin')
    for index, value in enumerate(my_excel):
        #exceltool.genfixedexcel(index, value,lubanOutputPath)
        generated_string = exceltool.genfixedexcel(index, value, lubanOutputPath)
    print('done!')
    #

    #打包工具需要
    #isClient = "True"
    folder_path =dir_path + r"\Luban\Luban.ClientServer\Templates\config\cs_unity_json"
    source_folder_class =dir_path + r"\CustomTem\class"
    source_folder_struct =dir_path + r"\CustomTem\struct"
    source_folder_self =dir_path + r"\CustomTem\server"

    if isClient =="True" :
        template.handle_template(source_folder_class,folder_path)
        os.system(dir_path + r"\Luban\gen_code_json.bat")
        print("client:已经生成class脚本和json数据")
        template.handle_template(source_folder_struct,folder_path)
        os.system(dir_path + r"\Luban\gen_code_json1.bat")
        print("client:已经生成blob脚本")
    else: 
        template.handle_template(source_folder_self,folder_path)
        os.system(dir_path + r"\Luban\gen_code_json_server.bat")
        print("server:已经生成json数据")
    
gen_base("False")

#input("Press Enter to exit...")









