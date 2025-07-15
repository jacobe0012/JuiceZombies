from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import shutil
import re



def createSkillNew(parent_dir,skilleffectNewName,skilleffectElementName):
    # 创建一个新的工作簿
    skillelewb = Workbook()
    
    skillelews = skillelewb.active
    # 将第一个工作表重命名为 "data"
    skillelews.title = "skill_effectElement|"
    
    skillelewslast_column1 = 1
    different_strings = ["id", "int", "all", "元素id"]  
    for idx, string in enumerate(different_strings, start=1):  
        cell = skillelews.cell(row=idx, column=skillelewslast_column1) 
        cell.value = string
        
    skillelewslast_column2 = 2
    different_strings = ["element_type", "int", "all", "元素类型"]  
    for idx, string in enumerate(different_strings, start=1):  
        cell = skillelews.cell(row=idx, column=skillelewslast_column2) 
        cell.value = string
        
    skillelewslast_column3 = 3
    different_strings = ["output_type", "int", "all", "输出类型"]  
    for idx, string in enumerate(different_strings, start=1):  
        cell = skillelews.cell(row=idx, column=skillelewslast_column3) 
        cell.value = string
        
    skillelewslast_column4 = 4
    different_strings = ["output_type_para", "array_int", "all", "输出类型参数值"]  
    for idx, string in enumerate(different_strings, start=1):  
        cell = skillelews.cell(row=idx, column=skillelewslast_column4) 
        cell.value = string
    skillelewslast_column5 = 5
    different_strings = ["bonus_type", "int", "all", "受属性影响类型"]  
    for idx, string in enumerate(different_strings, start=1):  
        cell = skillelews.cell(row=idx, column=skillelewslast_column5) 
        cell.value = string
        
        
        
    skillelewslast_column6 = 6
    different_strings = ["bonus_type_para", "array_int", "all", "受属性影响类型参数值"]  
    for idx, string in enumerate(different_strings, start=1):  
        cell = skillelews.cell(row=idx, column=skillelewslast_column6) 
        cell.value = string
        
    skillelewslast_column7 = 7
    different_strings = ["calc_type", "int", "all", "计算类型"]  
    for idx, string in enumerate(different_strings, start=1):  
        cell = skillelews.cell(row=idx, column=skillelewslast_column7) 
        cell.value = string
        
    skillelewslast_column8 = 8
    different_strings = ["calc_type_para", "array_int", "all", "计算类型参数值"]  
    for idx, string in enumerate(different_strings, start=1):  
        cell = skillelews.cell(row=idx, column=skillelewslast_column8) 
        cell.value = string
        
        
    skillelewslast_column9 = 9
    different_strings = ["state_type", "int", "all", "正负面类型"]  
    for idx, string in enumerate(different_strings, start=1):  
        cell = skillelews.cell(row=idx, column=skillelewslast_column9) 
        cell.value = string
        
    skillelewslast_column10 = 10
    different_strings = ["state_type_para", "array_int", "all", "正负面类型参数值"]  
    for idx, string in enumerate(different_strings, start=1):  
        cell = skillelews.cell(row=idx, column=skillelewslast_column10) 
        cell.value = string
        
        
    skillelewslast_column11 = 11
    different_strings = ["attr_id", "int", "all", "变更属性ID"]  
    for idx, string in enumerate(different_strings, start=1):  
        cell = skillelews.cell(row=idx, column=skillelewslast_column11) 
        cell.value = string
        
    skillelewslast_column12 = 12
    different_strings = ["attr_id_para", "array_int", "all", "变更属性ID参数值"]  
    for idx, string in enumerate(different_strings, start=1):  
        cell = skillelews.cell(row=idx, column=skillelewslast_column12) 
        cell.value = string
        
    skillelewslast_column13 = 13
    different_strings = ["control_type", "int", "all", "控制类型"]  
    for idx, string in enumerate(different_strings, start=1):  
        cell = skillelews.cell(row=idx, column=skillelewslast_column13) 
        cell.value = string
        
        
    skillelewslast_column14 = 14
    different_strings = ["control_type_para", "array_int", "all", "控制类型参数值"]  
    for idx, string in enumerate(different_strings, start=1):  
        cell = skillelews.cell(row=idx, column=skillelewslast_column14) 
        cell.value = string
        
    skillelewslast_column15 = 15
    different_strings = ["change_type", "int", "all", "替换类型"]  
    for idx, string in enumerate(different_strings, start=1):  
        cell = skillelews.cell(row=idx, column=skillelewslast_column15) 
        cell.value = string  
        
    skillelewslast_column16 = 16
    different_strings = ["change_type_para", "array_int", "all", "替换类型参数值"]  
    for idx, string in enumerate(different_strings, start=1):  
        cell = skillelews.cell(row=idx, column=skillelewslast_column16) 
        cell.value = string
        
    skillelewslast_column17 = 17
    different_strings = ["clear_type", "int", "all", "清除类型"]  
    for idx, string in enumerate(different_strings, start=1):  
        cell = skillelews.cell(row=idx, column=skillelewslast_column17) 
        cell.value = string  
        
    skillelewslast_column18 = 18
    different_strings = ["clear_type_para", "array_int", "all", "清除类型参数值"]  
    for idx, string in enumerate(different_strings, start=1):  
        cell = skillelews.cell(row=idx, column=skillelewslast_column18) 
        cell.value = string
        
        
    skillelewslast_column19 = 19
    different_strings = ["immune_type", "int", "all", "免疫类型"]  
    for idx, string in enumerate(different_strings, start=1):  
        cell = skillelews.cell(row=idx, column=skillelewslast_column19) 
        cell.value = string  
        
    skillelewslast_column20 = 20
    different_strings = ["immune_type_para", "array_int", "all", "免疫类型参数值"]  
    for idx, string in enumerate(different_strings, start=1):  
        cell = skillelews.cell(row=idx, column=skillelewslast_column20) 
        cell.value = string
        
        
    skillelewslast_column21 = 21
    different_strings = ["displace_from", "int", "all", "强制位移依据"]  
    for idx, string in enumerate(different_strings, start=1):  
        cell = skillelews.cell(row=idx, column=skillelewslast_column21) 
        cell.value = string  
        
    skillelewslast_column22 = 22
    different_strings = ["displace_from_para", "array_int", "all", "强制位移依据参数值"]  
    for idx, string in enumerate(different_strings, start=1):  
        cell = skillelews.cell(row=idx, column=skillelewslast_column22) 
        cell.value = string
        
    skillelewslast_column23 = 23
    different_strings = ["point_type", "int", "all", "选点类型"]  
    for idx, string in enumerate(different_strings, start=1):  
        cell = skillelews.cell(row=idx, column=skillelewslast_column23) 
        cell.value = string  
        
    skillelewslast_column24 = 24
    different_strings = ["point_type_para", "array_int", "all", "选点类型参数值"]  
    for idx, string in enumerate(different_strings, start=1):  
        cell = skillelews.cell(row=idx, column=skillelewslast_column24) 
        cell.value = string 
        
    skillelewslast_column25 = 25
    different_strings = ["power", "int", "all", "概率权重"]  
    for idx, string in enumerate(different_strings, start=1):  
        cell = skillelews.cell(row=idx, column=skillelewslast_column25) 
        cell.value = string
        
    skillelewslast_column26 = 26
    different_strings = ["pass_type", "int", "all", "穿墙类型"]  
    for idx, string in enumerate(different_strings, start=1):  
        cell = skillelews.cell(row=idx, column=skillelewslast_column26) 
        cell.value = string    
      
    skillelewslast_column27 = 27
    different_strings = ["bonus_other_type", "int", "all", "受其他影响类型"]  
    for idx, string in enumerate(different_strings, start=1):  
        cell = skillelews.cell(row=idx, column=skillelewslast_column27) 
        cell.value = string 
        
    skillelewslast_column28 = 28
    different_strings = ["bonus_other_type_para", "array_int", "all", "受其他影响类型参数"]  
    for idx, string in enumerate(different_strings, start=1):  
        cell = skillelews.cell(row=idx, column=skillelewslast_column28) 
        cell.value = string 

    skillelewslast_column29 = 29
    different_strings = ["direction_type", "int", "all", "方向类型"]  
    for idx, string in enumerate(different_strings, start=1):  
        cell = skillelews.cell(row=idx, column=skillelewslast_column29) 
        cell.value = string 
        
    skillelewslast_column30 = 30
    different_strings = ["direction_type_para", "array_int", "all", "方向类型参数"]  
    for idx, string in enumerate(different_strings, start=1):  
        cell = skillelews.cell(row=idx, column=skillelewslast_column30) 
        cell.value = string 

    skillelewslast_column31 = 31
    different_strings = ["hit_type", "int", "all", "方向类型"]  
    for idx, string in enumerate(different_strings, start=1):  
        cell = skillelews.cell(row=idx, column=skillelewslast_column31) 
        cell.value = string 
        
    skillelewslast_column32 = 32
    different_strings = ["hit_type_para", "array_int", "all", "方向类型参数"]  
    for idx, string in enumerate(different_strings, start=1):  
        cell = skillelews.cell(row=idx, column=skillelewslast_column32) 
        cell.value = string 

    skillelewslast_column33 = 33
    different_strings = ["pass_type_para", "array_int", "all", "穿墙类型参数"]  
    for idx, string in enumerate(different_strings, start=1):  
        cell = skillelews.cell(row=idx, column=skillelewslast_column33) 
        cell.value = string 
    # 指定原始文件路径和目标文件路径
    olddir= parent_dir + r"\config\battle\skill_effect.xlsx"  
    # 复制原始文件到目标文件路径
    shutil.copy(olddir, skilleffectNewName)
    
    effectwb = load_workbook(skilleffectNewName, data_only=True)
    
    
    sheet_names = effectwb.sheetnames
    # 遍历工作表名称，找到包含"data"的工作表
    matching_sheets = [sheet_name for sheet_name in sheet_names if 'skill_effect|' in sheet_name.lower()]
    effectws =effectwb[matching_sheets[0]] 
    effectws.title = "skill_effectNew|"
    # 删除除目标表之外的其他表
    for sheet_name in sheet_names:
        if '|' not in sheet_name:
            sheet = effectwb[sheet_name]
            effectwb.remove(sheet)

    effect_string_idx = None

    id =None
    for cell in effectws[1]:
        if cell.value == "effect_string":
            effect_string_idx = cell.column
        elif cell.value == "id":
            id = cell.column           

    last_column1 = effect_string_idx + 1
    different_strings = ["trigger_type", "int", "all", "触发效果类型"]  
    for idx, string in enumerate(different_strings, start=1):  
        cell = effectws.cell(row=idx, column=effect_string_idx + 1) 
        cell.value = string

    last_column2 = effect_string_idx + 2
    different_strings = ["trigger_type_para", "array_int", "all", "触发效果类型参数"]  
    for idx, string in enumerate(different_strings, start=1):  
        cell = effectws.cell(row=idx, column=effect_string_idx + 2) 
        cell.value = string

    last_column3 = effect_string_idx + 3
    different_strings = ["condition_type", "int", "all", "条件类型"]  
    for idx, string in enumerate(different_strings, start=1):  
        cell = effectws.cell(row=idx, column=effect_string_idx + 3) 
        cell.value = string 
    
    last_column4 = effect_string_idx + 4
    different_strings = ["condition_type_para", "array_int", "all", "条件类型参数"]  
    for idx, string in enumerate(different_strings, start=1):  
        cell = effectws.cell(row=idx, column=effect_string_idx + 4) 
        cell.value = string 
              
    last_column5 = effect_string_idx + 5
    different_strings = ["compare_type", "int", "all", "比较类型"]  
    for idx, string in enumerate(different_strings, start=1):  
        cell = effectws.cell(row=idx, column=effect_string_idx + 5) 
        cell.value = string 
         
    last_column6 = effect_string_idx + 6
    different_strings = ["compare_type_para", "array_int", "all", "比较类型参数"]  
    for idx, string in enumerate(different_strings, start=1):  
        cell = effectws.cell(row=idx, column=effect_string_idx + 6) 
        cell.value = string 
        
        
    last_column7 = effect_string_idx + 7
    different_strings = ["delay_type", "int", "all", "延迟类型"]  
    for idx, string in enumerate(different_strings, start=1):  
        cell = effectws.cell(row=idx, column=effect_string_idx + 7) 
        cell.value = string 
         
    last_column8 = effect_string_idx + 8
    different_strings = ["delay_type_para", "array_int", "all", "延迟类型参数"]  
    for idx, string in enumerate(different_strings, start=1):  
        cell = effectws.cell(row=idx, column=effect_string_idx + 8) 
        cell.value = string         
        
    last_column9 = effect_string_idx + 9
    different_strings = ["calc_type", "int", "all", "计算类型"]  
    for idx, string in enumerate(different_strings, start=1):  
        cell = effectws.cell(row=idx, column=effect_string_idx + 9) 
        cell.value = string 
         
    last_column10 = effect_string_idx + 10
    different_strings = ["calc_type_para", "array_int", "all", "计算类型参数 "]  
    for idx, string in enumerate(different_strings, start=1):  
        cell = effectws.cell(row=idx, column=effect_string_idx + 10) 
        cell.value = string      

    last_column11 = effect_string_idx + 11
    different_strings = ["extra_type", "int", "all", "额外类型"]  
    for idx, string in enumerate(different_strings, start=1):  
        cell = effectws.cell(row=idx, column=effect_string_idx + 11) 
        cell.value = string 
         
    last_column12 = effect_string_idx + 12
    different_strings = ["extra_type_para", "array_int", "all", "额外类型参数 "]  
    for idx, string in enumerate(different_strings, start=1):  
        cell = effectws.cell(row=idx, column=effect_string_idx + 12) 
        cell.value = string      

    #非触发器类型
    last_column13 = effect_string_idx + 13
    different_strings = ["search_type", "int", "all", "索敌类型"]  
    for idx, string in enumerate(different_strings, start=1):  
        cell = effectws.cell(row=idx, column=effect_string_idx + 13) 
        cell.value = string 
         
    last_column14 = effect_string_idx + 14
    different_strings = ["search_type_para", "array_int", "all", "索敌类型参数 "]  
    for idx, string in enumerate(different_strings, start=1):  
        cell = effectws.cell(row=idx, column=effect_string_idx + 14) 
        cell.value = string           
        
    last_column15 = effect_string_idx + 15
    different_strings = ["range_type", "int", "all", "范围类型"]  
    for idx, string in enumerate(different_strings, start=1):  
        cell = effectws.cell(row=idx, column=effect_string_idx + 15) 
        cell.value = string 
         
    last_column16 = effect_string_idx + 16
    different_strings = ["range_type_para", "array_int", "all", "范围类型参数 "]  
    for idx, string in enumerate(different_strings, start=1):  
        cell = effectws.cell(row=idx, column=effect_string_idx + 16) 
        cell.value = string 

    last_column17 = effect_string_idx + 17
    different_strings = ["deviate_type", "int", "all", "偏移类型"]  
    for idx, string in enumerate(different_strings, start=1):  
        cell = effectws.cell(row=idx, column=effect_string_idx + 17) 
        cell.value = string 
         
    last_column18 = effect_string_idx + 18
    different_strings = ["deviate_type_para", "array_int", "all", "偏移类型参数 "]  
    for idx, string in enumerate(different_strings, start=1):  
        cell = effectws.cell(row=idx, column=effect_string_idx + 18) 
        cell.value = string 
    
    last_column19 = effect_string_idx + 19
    different_strings = ["target", "int", "all", "作用对象"]  
    for idx, string in enumerate(different_strings, start=1):  
        cell = effectws.cell(row=idx, column=effect_string_idx + 19) 
        cell.value = string 
         
    last_column20 = effect_string_idx + 20
    different_strings = ["target_para", "array_int", "all", "作用对象参数 "]  
    for idx, string in enumerate(different_strings, start=1):  
        cell = effectws.cell(row=idx, column=effect_string_idx + 20) 
        cell.value = string 
        
    last_column21 = effect_string_idx + 21
    different_strings = ["element_trigger", "array_int", "all", "自加的元素类触发器 "]  
    for idx, string in enumerate(different_strings, start=1):  
        cell = effectws.cell(row=idx, column=effect_string_idx + 21) 
        cell.value = string 
    
    last_column22 = effect_string_idx + 22
    different_strings = ["element_list", "array_int", "all", "元素列表"]  
    for idx, string in enumerate(different_strings, start=1):  
        cell = effectws.cell(row=idx, column=effect_string_idx + 22) 
        cell.value = string 
    last_column23 = effect_string_idx + 23
    different_strings = ["power", "int", "all", "概率权重"]  
    for idx, string in enumerate(different_strings, start=1):  
        cell = effectws.cell(row=idx, column=effect_string_idx + 23) 
        cell.value = string  

    last_column24 = effect_string_idx + 24
    different_strings = ["effect_type", "int", "all", "效果类型"]  
    for idx, string in enumerate(different_strings, start=1):  
        cell = effectws.cell(row=idx, column=effect_string_idx + 24) 
        cell.value = string      
        
    last_column25 = effect_string_idx + 25
    different_strings = ["attr_id", "int", "all", "变更属性ID"]  
    for idx, string in enumerate(different_strings, start=1):  
        cell = effectws.cell(row=idx, column=effect_string_idx + 25) 
        cell.value = string  

    last_column26 = effect_string_idx + 26
    different_strings = ["attr_id_para", "array_int", "all", "变更属性ID参数值"]  
    for idx, string in enumerate(different_strings, start=1):  
        cell = effectws.cell(row=idx, column=effect_string_idx + 26) 
        cell.value = string   
        
    last_column27 = effect_string_idx + 27
    different_strings = ["target_lock_on", "int", "all", "索敌目标"]  
    for idx, string in enumerate(different_strings, start=1):  
        cell = effectws.cell(row=idx, column=effect_string_idx + 27) 
        cell.value = string 
         
    last_column28 = effect_string_idx + 28
    different_strings = ["target_lock_on_para", "array_int", "all", "索敌目标参数 "]  
    for idx, string in enumerate(different_strings, start=1):  
        cell = effectws.cell(row=idx, column=effect_string_idx + 28) 
        cell.value = string 
        
    elementId = 0
    if effect_string_idx is not None:
    # 从该列的第5行开始遍历到最后一行
        for row in range(5, effectws.max_row+1):
            effect_str =effectws.cell(row=row, column=effect_string_idx).value
            if effect_str =="" or effect_str ==None :
               continue
            print(f"已转换skill_effect的第:{row}行")       
            power = re.search(r'power=(\d+)', effect_str).group(1)
            if power == "" or power ==None:
                power ="N/A"
            effectws.cell(row=row, column=last_column23,value=str(power))
            
            effect_type = re.search(r'effect_type=(\d+)', effect_str).group(1)
            if effect_type == "" or effect_type ==None:
                effect_type ="N/A"
            effectws.cell(row=row, column=last_column24,value=str(effect_type))
            
            if str(effect_type) == "2" :
                
                # 使用正则表达式提取 trigger_type 值
                trigger_type = re.search(r'trigger_type=(\d+)', effect_str).group(1)
                trigger_type_para = re.search(r'trigger_type_para=\[(.*?)\]', effect_str).group(1)

                condition_type = re.search(r'condition_type=(\d+)', effect_str).group(1)
                condition_type_para = re.search(r'condition_type_para=\[(.*?)\]', effect_str).group(1)
                
                compare_type = re.search(r'compare_type=(\d+)', effect_str).group(1)
                compare_type_para = re.search(r'compare_type_para=\[(.*?)\]', effect_str).group(1)
                
                delay_type = re.search(r'delay_type=(\d+)', effect_str).group(1)
                delay_type_para = re.search(r'delay_type_para=\[(.*?)\]', effect_str).group(1)
                
                calc_type = re.search(r'calc_type=(\d+)', effect_str).group(1)
                calc_type_para = re.search(r'calc_type_para=\[(.*?)\]', effect_str).group(1)
                
                extra_type = re.search(r'extra_type=(\d+)', effect_str).group(1)
                extra_type_para = re.search(r'extra_type_para=\[(.*?)\]', effect_str).group(1)

                if trigger_type == "" or trigger_type ==None:
                    trigger_type ="N/A"
                if trigger_type_para == "" or trigger_type_para ==None:
                    trigger_type_para ="N/A"
                if condition_type == "" or condition_type ==None:
                    condition_type ="N/A"
                if condition_type_para == "" or condition_type_para ==None:
                    condition_type_para ="N/A"    
                if compare_type == "" or compare_type ==None:
                    compare_type ="N/A"
                if compare_type_para == "" or compare_type_para ==None:
                    compare_type_para ="N/A"
                if delay_type == "" or delay_type ==None:
                    delay_type ="N/A"
                if delay_type_para == "" or delay_type_para ==None:
                    delay_type_para ="N/A" 
                if calc_type == "" or calc_type ==None:
                    calc_type ="N/A"
                if calc_type_para == "" or calc_type_para ==None:
                    calc_type_para ="N/A"               
                    
                if extra_type == "" or extra_type ==None:
                    extra_type ="N/A"
                if extra_type_para == "" or extra_type_para ==None:
                    extra_type_para ="N/A" 
                
                    
                effectws.cell(row=row, column=last_column1,value=str(trigger_type))
                effectws.cell(row=row, column=last_column2,value=str(trigger_type_para))
                
                effectws.cell(row=row, column=last_column3,value=str(condition_type))
                effectws.cell(row=row, column=last_column4,value=str(condition_type_para))    

                effectws.cell(row=row, column=last_column5,value=str(compare_type))
                effectws.cell(row=row, column=last_column6,value=str(compare_type_para))
                
                effectws.cell(row=row, column=last_column7,value=str(delay_type))
                effectws.cell(row=row, column=last_column8,value=str(delay_type_para))           

                effectws.cell(row=row, column=last_column9,value=str(calc_type))
                effectws.cell(row=row, column=last_column10,value=str(calc_type_para))  

                effectws.cell(row=row, column=last_column11,value=str(extra_type))
                effectws.cell(row=row, column=last_column12,value=str(extra_type_para))  

                #print(effectws.cell(row=row, column=last_column3).value)
                
                if  str(effectws.cell(row=row, column=last_column3).value)  == "2":
                    element_trigger_before=str(effectws.cell(row=row, column = last_column4).value).split(";")
                    idstr = str(effectws.cell(row=row, column=id).value) 
                    for item in element_trigger_before:
                        for row0 in range(5, effectws.max_row+1):
                            if str(effectws.cell(row=row0, column=id).value) ==item:
                                effectws.cell(row=row0, column=last_column21).value = idstr
                        


                #last_column3    
            elif str(effect_type) == "1" :
                
                element_values = []
                #非触发器类型
                # 定义要匹配的模式
                
                element_1 = re.compile(r'element_1:{.*?};', re.DOTALL)
                element_2 = re.compile(r'element_2:{.*?};', re.DOTALL)
                element_3 = re.compile(r'element_3:{.*?};', re.DOTALL)
                # 使用 re.search 进行匹配
                elementstr =[]
                elementstr.append(element_1.search(effect_str))
                elementstr.append(element_2.search(effect_str))
                elementstr.append(element_3.search(effect_str))
                for item in elementstr :
                    if item:
                        itemcontent = item.group(0)
                        elementId += 1     
                        skillelews.cell(row=elementId+4, column=1,value=str(elementId))
                        
                        skillele_search=[]
                        patterns = [
                            (r'element_type=(\d+)', 'element_type'),
                            (r'output_type=(\d+)', 'output_type'),
                            (r'output_type_para=\[(.*?)\]', 'output_type_para'),
                            (r'bonus_type=(\d+)', 'bonus_type'),
                            (r'bonus_type_para=\[(.*?)\]', 'bonus_type_para'),
                            (r'calc_type=(\d+)', 'calc_type'),
                            (r'calc_type_para=\[(.*?)\]', 'calc_type_para'),
                            (r'state_type=(\d+)', 'state_type'),
                            (r'state_type_para=\[(.*?)\]', 'state_type_para'),
                            (r'attr_id=(\d+)', 'attr_id'),
                            (r'attr_id_para=\[(.*?)\]', 'attr_id_para'),
                            (r'control_type=(\d+)', 'control_type'),
                            (r'control_type_para=\[(.*?)\]', 'control_type_para'),
                            (r'change_type=(\d+)', 'change_type'),
                            (r'change_type_para=\[(.*?)\]', 'change_type_para'),
                            (r'clear_type=(\d+)', 'clear_type'),
                            (r'clear_type_para=\[(.*?)\]', 'clear_type_para'),
                            (r'immune_type=(\d+)', 'immune_type'),
                            (r'immune_type_para=\[(.*?)\]', 'immune_type_para'),
                            (r'displace_from=(\d+)', 'displace_from'),
                            (r'displace_from_para=\[(.*?)\]', 'displace_from_para'),
                            (r'point_type=(\d+)', 'point_type'),
                            (r'point_type_para=\[(.*?)\]', 'point_type_para'),
                            (r'power=(\d+)', 'power'),
                            (r'pass_type=(\d+)', 'pass_type'),
                            (r'bonus_other_type=(\d+)', 'bonus_other_type'),
                            (r'bonus_other_type_para=\[(.*?)\]', 'bonus_other_type_para'),
                            (r'direction_type=(\d+)', 'direction_type'),
                            (r'direction_type_para=\[(.*?)\]', 'direction_type_para'),
                            (r'hit_type=(\d+)', 'hit_type'),
                            (r'hit_type_para=\[(.*?)\]', 'hit_type_para'),
                            (r'pass_type_para=\[(.*?)\]', 'pass_type_para'),
                            
                        ]
                        # 对每个正则表达式模式进行搜索和处理
                        for pattern, field_name in patterns:
                            match = re.search(pattern, itemcontent)
                            if match:
                                skillele_search.append(match.group(1))
                            else:
                                skillele_search.append("")
                                   
                        for index, value in enumerate(skillele_search):
                            if value != "" and value !=None:            
                                columnindex = index+2
                                #print(f"The element '{value}' is at index {index}.")
                                
                                skillelews.cell(row=elementId+4, column=columnindex,value=str(value))
                                
                                #skillelewscell = skillelews.cell(row=elementId+4, column=columnindex) 
                                #skillelewscell.value = str(value)
                                

                        element_values.append(elementId)
                        #print("itemcontent 内容:")
                        #print(itemcontent)
                element_str_values = [str(value) for value in element_values]
                merged_value = ';'.join(element_str_values)
                effectws.cell(row=row, column=last_column22,value=str(merged_value)) 
                

                #skillelews.cell(row=1, column=1,value=str("value"))
                search_type = re.search(r'search_type=(\d+)', effect_str).group(1)
                search_type_para = re.search(r'search_type_para=\[(.*?)\]', effect_str).group(1)

                range_type = re.search(r'range_type=(\d+)', effect_str).group(1)
                range_type_para = re.search(r'range_type_para=\[(.*?)\]', effect_str).group(1)
                
                deviate_type = re.search(r'deviate_type=(\d+)', effect_str).group(1)
                deviate_type_para = re.search(r'deviate_type_para=\[(.*?)\]', effect_str).group(1)

                target = re.search(r'target=(\d+)', effect_str).group(1)
                target_para = re.search(r'target_para=\[(.*?)\]', effect_str).group(1)   

                condition_type = re.search(r'condition_type=(\d+)', effect_str).group(1)
                condition_type_para = re.search(r'condition_type_para=\[(.*?)\]', effect_str).group(1)
                
                target_lock_on = re.search(r'target_lock_on=(\d+)', effect_str).group(1)
                target_lock_on_para = re.search(r'target_lock_on_para=\[(.*?)\]', effect_str).group(1)   
                if search_type == "" or search_type ==None:
                    search_type ="N/A"
                if search_type_para == "" or search_type_para ==None:
                    search_type_para ="N/A"
                if range_type == "" or range_type ==None:
                    range_type ="N/A"
                if range_type_para == "" or range_type_para ==None:
                    range_type_para ="N/A"    
                if deviate_type == "" or deviate_type ==None:
                    deviate_type ="N/A"
                if deviate_type_para == "" or deviate_type_para ==None:
                    deviate_type_para ="N/A"
                if target == "" or target ==None:
                    target ="N/A"
                if target_para == "" or target_para ==None:
                    target_para ="N/A" 
                if condition_type == "" or condition_type ==None:
                    condition_type ="N/A"
                if condition_type_para == "" or condition_type_para ==None:
                    condition_type_para ="N/A"    
                if target_lock_on == "" or target_lock_on ==None:
                    target_lock_on ="N/A"
                if target_lock_on_para == "" or target_lock_on_para ==None:
                    target_lock_on_para ="N/A" 
                    
                    
                effectws.cell(row=row, column=last_column3,value=str(condition_type))
                effectws.cell(row=row, column=last_column4,value=str(condition_type_para))    

                effectws.cell(row=row, column=last_column13,value=str(search_type))
                effectws.cell(row=row, column=last_column14,value=str(search_type_para))    

                effectws.cell(row=row, column=last_column15,value=str(range_type))
                effectws.cell(row=row, column=last_column16,value=str(range_type_para))
                
                effectws.cell(row=row, column=last_column17,value=str(deviate_type))
                effectws.cell(row=row, column=last_column18,value=str(deviate_type_para))           

                effectws.cell(row=row, column=last_column19,value=str(target))
                effectws.cell(row=row, column=last_column20,value=str(target_para)) 

                effectws.cell(row=row, column=last_column27,value=str(target_lock_on))
                effectws.cell(row=row, column=last_column28,value=str(target_lock_on_para)) 

            elif str(effect_type) == "3" :
                # 使用正则表达式提取 trigger_type 值
                attr_id_ser = re.search(r'attr_id=(\d+)', effect_str).group(1)
                attr_id_para_ser = re.search(r'attr_id_para=\[(.*?)\]', effect_str).group(1)
                if attr_id_ser == "" or attr_id_ser ==None:
                    attr_id_ser ="N/A"
                if attr_id_para_ser == "" or attr_id_para_ser ==None:
                    attr_id_para_ser ="N/A"
                effectws.cell(row=row, column=last_column25,value=str(attr_id_ser))
                effectws.cell(row=row, column=last_column26,value=str(attr_id_para_ser))
                
                
    skillelewb.save(skilleffectElementName)    
    skillelewb.close() 
    # 遍历工作表中的每一行，删除指定列（从后往前）
    effectws.delete_cols(effect_string_idx) 
    effectwb.save(skilleffectNewName)    
    effectwb.close() 
 